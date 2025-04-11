import yfinance as yf
import pandas as pd
import os
from openpyxl  import load_workbook, Workbook
from openpyxl.utils.dataframe  import dataframe_to_rows

# 조회할 ETF 티커 리스트
tickers = ["SPY","qqq","IWD","IWM","IWN","MTUM","EFA","EWJ","EEM","VGK", "TLT","IEF","SHY",  "VNQ",  "LQD", "DBC","VNQ","BWX", "AGG","GLD", "DBC","BIL","HYG","REM",   "EDV", "LTPZ", "LQD", "EMLC"]
# tickers = "SPY"
file_name="stock_data.xlsx"
sheet_name = "B"


# 세션 생성 후 다운로드 시도
session = yf.download(tickers, period="1d")
# 데이터 가져오기 (최근 1년치 종가)
data = yf.download(tickers, period="1y")["Close"]

df=pd.DataFrame(data)

# 엑셀 파일 불러오기
book = load_workbook(file_name)
sheet = book[sheet_name]

# 마지막 열 찾기
last_raw = sheet.max_row + 1

df.reset_index()
# df.style.hide_index()
# 엑셀 파일 쓰기
for r in dataframe_to_rows(df,index = True, header = True):
    sheet.append(r)
book.save(file_name)
